"""
Crawler Web Server
==================
Serves the frontend UI and runs the crawl4ai crawler on demand.
Supports two modes:
  - Sitemap: parse sitemap.xml and crawl listed URLs
  - Spider:  start from a root URL and follow internal links

Anti-bot strategy (when ScrapingBee API key is provided):
  1. Try direct crawl with Crawl4AI first (free, fast)
  2. If bot-detection phrases are found in the response, automatically
     retry via ScrapingBee (paid, bypasses protection) — saves credits
     by only using ScrapingBee where actually needed.

Requirements:
    pip install crawl4ai aiohttp lxml flask openpyxl

Run:
    python server.py
Then open: http://localhost:5000
"""

import asyncio
import io
import os
import re
import threading
import uuid
from urllib.parse import urlparse

import aiohttp
from flask import Flask, jsonify, request, send_file
from lxml import etree
from lxml import html as lhtml
from crawl4ai import AsyncWebCrawler, CrawlerRunConfig, CacheMode
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

# In-memory job store
jobs = {}

# Any ONE of these phrases in the response body triggers the ScrapingBee fallback
BOT_DETECTION_PHRASES = [
    "temporary error. please try again.",   # exact string confirmed by user
    "solve a puzzle",
    "before proceeding to your request",
    "captcha",
    "enable javascript",
    "ddos protection",
    "checking your browser",
    "ray id",
    "datadome",
    "robot or human",
    "access denied",
]


# URL path segments and extensions to skip entirely
SKIP_EXTENSIONS = {'.css', '.js', '.png', '.jpg', '.jpeg', '.gif',
                   '.svg', '.ico', '.woff', '.woff2', '.ttf', '.eot', '.mp4',
                   '.mp3', '.zip', '.xml'}
# Note: .pdf intentionally excluded — PDFs are handled by the Claude API (tier 4)
SKIP_PATH_SEGMENTS = {'/css/', '/js/', '/fonts/', '/images/', '/img/',
                      '/static/', '/assets/', '/media/', '/file%20library/',
                      '/style%20library/'}


def should_skip_url(url: str) -> bool:
    """Return True if the URL points to a non-page asset we don't want to crawl."""
    lower = url.lower().split('?')[0]   # ignore query string for extension check
    # Skip by file extension
    if any(lower.endswith(ext) for ext in SKIP_EXTENSIONS):
        return True
    # Skip by path segment
    if any(seg in lower for seg in SKIP_PATH_SEGMENTS):
        return True
    return False


# ──────────────────────────────────────────────
# PDF detection
# ──────────────────────────────────────────────

def is_pdf_url(url: str) -> bool:
    """Return True if the URL clearly points to a PDF file."""
    return url.lower().split('?')[0].endswith('.pdf')


# ──────────────────────────────────────────────
# Bot detection check
# ──────────────────────────────────────────────

def is_blocked(html: str) -> bool:
    """Return True if the page looks like a bot-detection wall.
    Any single matching phrase is enough to trigger the fallback."""
    if not html:
        return True
    sample = html[:10000].lower()
    return any(phrase in sample for phrase in BOT_DETECTION_PHRASES)


# ──────────────────────────────────────────────
# ScrapingBee fallback
# ──────────────────────────────────────────────

async def fetch_via_scrapingbee(url: str, api_key: str,
                                session: aiohttp.ClientSession,
                                stealth: bool = False) -> str:
    """
    Fetch a URL through ScrapingBee's API.
    stealth=False  → premium_proxy  (~5 credits/request)
    stealth=True   → stealth_proxy  (~75 credits/request, most powerful)
    Returns raw HTML string, raises RuntimeError on failure.
    """
    params = {
        "api_key": api_key,
        "url": url,
        "render_js": "true",
        "block_ads": "true",
        "block_resources": "false",
    }
    if stealth:
        params["stealth_proxy"] = "true"
    else:
        params["premium_proxy"] = "true"

    endpoint = "https://app.scrapingbee.com/api/v1/"
    try:
        async with session.get(
            endpoint, params=params,
            timeout=aiohttp.ClientTimeout(total=90)
        ) as resp:
            if resp.status == 200:
                return await resp.text()
            else:
                text = await resp.text()
                raise RuntimeError(f"ScrapingBee HTTP {resp.status}: {text[:200]}")
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"ScrapingBee request failed: {e}")


# ──────────────────────────────────────────────
# Claude API fallback (Tier 4 — last resort)
# ──────────────────────────────────────────────

async def fetch_via_claude(url: str, fields: list,
                           claude_api_url: str,
                           session: aiohttp.ClientSession) -> dict:
    """
    Fetch page content via the Claude-based extraction API.
    Works for both web pages and PDF files.
    Maps API response: text -> content, page_title -> title/h1, image -> image_url
    """
    # Always request page_title; also request image if image_url field is selected
    api_fields = [
        {"name": "page_title", "description": "The name or title of this page or document",
         "type": "string", "required": True},
    ]
    if "image_url" in fields:
        api_fields.append({
            "name": "image",
            "description": "The URL of the main graphic or hero image on this page",
            "type": "array",
            "required": False
        })

    payload = {"url": url, "fields": api_fields}
    try:
        async with session.post(
            claude_api_url,
            json=payload,
            headers={"Content-Type": "application/json"},
            timeout=aiohttp.ClientTimeout(total=60)
        ) as resp:
            if resp.status != 200:
                raise RuntimeError(f"Claude API HTTP {resp.status}")
            data = await resp.json()

            # Extract page_title — may be top-level or nested under "fields"
            nested = data.get("fields", {}) or {}
            page_title = data.get("page_title") or nested.get("page_title", "")

            # Extract image — API returns an array, take first entry
            raw_image = data.get("image") or nested.get("image", [])
            if isinstance(raw_image, list):
                image_url = raw_image[0] if raw_image else ""
            else:
                image_url = str(raw_image) if raw_image else ""

            content = data.get("text", "")

            return {
                "url": url,
                "content": content,
                "title": page_title,
                "h1": page_title,
                "image_url": image_url,
                "meta_description": "",
                "word_count": str(len(content.split())) if content else "0",
                "links": "",
            }
    except Exception as e:
        raise RuntimeError(f"Claude API request failed: {e}")


# ──────────────────────────────────────────────
# Unified page fetcher (direct → ScrapingBee fallback)
# ──────────────────────────────────────────────

class PageResult:
    """Minimal result object compatible with extract_field()."""
    def __init__(self, url, html, metadata=None, links=None, success=True, error_message=""):
        self.url = url
        self.html = html
        self.markdown = ""          # not used — we extract from html directly
        self.metadata = metadata or {}
        self.links = links or {}
        self.success = success
        self.error_message = error_message


async def fetch_page(url: str, crawler: AsyncWebCrawler, config: CrawlerRunConfig,
                     api_key: str, session: aiohttp.ClientSession, job: dict,
                     claude_api_url: str = "", fields: list = None) -> PageResult:
    """
    Four-tier fetch strategy:
      1. Direct crawl via Crawl4AI         (free)
      2. ScrapingBee with premium_proxy    (~5 credits)
      3. ScrapingBee with stealth_proxy    (~75 credits)
      4. Claude extraction API             (last resort — no HTML, content+title only)
    """
    if fields is None:
        fields = []
    def make_result(html):
        title_m = re.search(r'<title[^>]*>(.*?)</title>', html, re.I | re.S)
        title = re.sub(r'<[^>]+>', '', title_m.group(1)).strip() if title_m else ""
        return PageResult(url=url, html=html, metadata={"title": title})

    # PDFs: skip tiers 1-3 entirely, go straight to Claude API
    if is_pdf_url(url):
        if claude_api_url:
            pass  # fall through to tier 4 below
        else:
            return PageResult(url=url, html="", success=False,
                              error_message="PDF skipped — no Claude API URL configured")

    # Tier 1: direct crawl (skipped for PDFs)
    elif True:
        try:
            result = await crawler.arun(url=url, config=config)
            if result.success and not is_blocked(result.html):
                return result
        except Exception:
            pass

    # Tiers 2 & 3: ScrapingBee (skip for PDFs — Claude API handles them better)
    if api_key and not is_pdf_url(url):
        # Tier 2: premium proxy
        job["log"].append(f"  ↳ Blocked — trying ScrapingBee premium proxy: {url}")
        try:
            html = await fetch_via_scrapingbee(url, api_key, session, stealth=False)
            if html and not is_blocked(html):
                job["log"].append(f"OK {url} (via ScrapingBee premium proxy)")
                return make_result(html)
            job["log"].append(f"  ↳ Premium proxy blocked — escalating to stealth proxy: {url}")
        except Exception as e:
            job["log"].append(f"  ↳ Premium proxy error ({e}) — escalating to stealth proxy: {url}")

        # Tier 3: stealth proxy
        try:
            html = await fetch_via_scrapingbee(url, api_key, session, stealth=True)
            if html and not is_blocked(html):
                job["log"].append(f"OK {url} (via ScrapingBee stealth proxy)")
                return make_result(html)
            job["log"].append(f"  ↳ Stealth proxy blocked — escalating to Claude API: {url}")
        except Exception as e:
            job["log"].append(f"  ↳ Stealth proxy error ({e}) — escalating to Claude API: {url}")
    elif not is_pdf_url(url):
        job["log"].append(f"  ↳ Blocked — no ScrapingBee key, trying Claude API: {url}")

    # Tier 4: Claude extraction API
    if claude_api_url:
        try:
            job["log"].append(f"  ↳ Trying Claude API: {url}")
            extracted = await fetch_via_claude(url, fields, claude_api_url, session)
            job["log"].append(f"OK {url} (via Claude API)")
            # Return a PageResult with pre-extracted fields stored in metadata
            result = PageResult(url=url, html="", metadata={"title": extracted.get("title", "")})
            result._extracted = extracted   # stash for extract_field to use
            return result
        except Exception as e:
            job["log"].append(f"  ↳ Claude API error: {e}")

    return PageResult(url=url, html="", success=False,
                      error_message="BLOCKED: All four fetch tiers failed")


# ──────────────────────────────────────────────
# Sitemap fetching
# ──────────────────────────────────────────────

async def fetch_sitemap_urls(sitemap_url: str, session: aiohttp.ClientSession) -> list:
    headers = {"User-Agent": "Mozilla/5.0 (compatible; SitemapCrawler/1.0)"}
    try:
        async with session.get(sitemap_url, headers=headers,
                               timeout=aiohttp.ClientTimeout(total=30)) as resp:
            resp.raise_for_status()
            content = await resp.read()
    except Exception as e:
        raise RuntimeError(f"Could not fetch sitemap: {e}")

    try:
        root = etree.fromstring(content)
    except etree.XMLSyntaxError as e:
        raise RuntimeError(f"Could not parse sitemap XML: {e}")

    ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
    urls = []

    child_sitemaps = root.findall(".//sm:sitemap/sm:loc", ns)
    if child_sitemaps:
        for loc_el in child_sitemaps:
            child_urls = await fetch_sitemap_urls(loc_el.text.strip(), session)
            urls.extend(child_urls)
    else:
        page_urls = root.findall(".//sm:url/sm:loc", ns)
        urls = [el.text.strip() for el in page_urls if el.text and not should_skip_url(el.text.strip())]

    return urls


# ──────────────────────────────────────────────
# Spider: discover URLs by following internal links
# ──────────────────────────────────────────────

def extract_internal_links(html_content: str, base_url: str) -> list:
    base_domain = urlparse(base_url).netloc
    found = []
    try:
        doc = lhtml.fromstring(html_content)
        doc.make_links_absolute(base_url)
        for el, attr, link, _ in doc.iterlinks():
            if attr != 'href':
                continue
            parsed = urlparse(link)
            if parsed.scheme in ('http', 'https') and parsed.netloc == base_domain:
                clean = link.split('#')[0].rstrip('/')
                if clean and not should_skip_url(clean):
                    found.append(clean)
    except Exception:
        pass
    return list(set(found))


async def spider_crawl(job_id: str, start_url: str, fields: list,
                       max_pages: int, api_key: str, claude_api_url: str = "") -> list:
    job = jobs[job_id]
    job["status"] = "crawling"
    job["log"] = [f"Spider starting at: {start_url}"]
    if api_key:
        job["log"].append("ScrapingBee fallback ENABLED (premium → stealth escalation)")
    else:
        job["log"].append("WARNING: No ScrapingBee API key — blocked pages will try Claude API only")
    if claude_api_url:
        job["log"].append("Claude API fallback ENABLED (tier 4 last resort)")

    config = CrawlerRunConfig(
        cache_mode=CacheMode.BYPASS,
        word_count_threshold=10,
        page_timeout=30000,
    )

    visited = set()
    queue = [start_url.rstrip('/')]
    rows = []
    semaphore = asyncio.Semaphore(5)

    async with AsyncWebCrawler() as crawler:
        async with aiohttp.ClientSession() as session:

            async def crawl_one(url):
                async with semaphore:
                    try:
                        result = await fetch_page(url, crawler, config, api_key, session, job, claude_api_url, fields)
                        if not result.success:
                            job["log"].append(f"SKIP {url}: {result.error_message}")
                            job["progress"] = job.get("progress", 0) + 1
                            return None, []
                        row = {f: extract_field(f, result) for f in fields}
                        new_links = extract_internal_links(result.html, url) if result.html else []
                        job["progress"] = job.get("progress", 0) + 1
                        job["log"].append(f"OK {url}")
                        return row, new_links
                    except Exception as e:
                        job["progress"] = job.get("progress", 0) + 1
                        job["log"].append(f"ERR {url}: {e}")
                        return None, []

            while queue and len(visited) < max_pages:
                batch = []
                while queue and len(batch) < 10:
                    url = queue.pop(0)
                    if url not in visited:
                        visited.add(url)
                        batch.append(url)

                if not batch:
                    break

                job["total"] = len(visited) + len(queue)
                tasks = [crawl_one(url) for url in batch]
                results = await asyncio.gather(*tasks)

                for row, new_links in results:
                    if row:
                        rows.append(row)
                    for link in new_links:
                        clean = link.rstrip('/')
                        if clean not in visited and clean not in queue:
                            queue.append(clean)

                job["total"] = len(visited) + len(queue)

    job["log"].append(f"Spider complete. {len(rows)} pages crawled.")
    return rows


# ──────────────────────────────────────────────
# Content extraction helpers
# ──────────────────────────────────────────────

def extract_hero_image(html: str) -> str:
    og = re.search(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\'](https?://[^"\']+)["\']', html, re.I)
    if og:
        return og.group(1)
    og2 = re.search(r'<meta[^>]+content=["\'](https?://[^"\']+)["\'][^>]+property=["\']og:image["\']', html, re.I)
    if og2:
        return og2.group(1)
    tw = re.search(r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\'](https?://[^"\']+)["\']', html, re.I)
    if tw:
        return tw.group(1)
    imgs = re.findall(r'<img[^>]+src=["\'](https?://[^"\']+)["\']', html, re.I)
    for src in imgs:
        lower = src.lower()
        if any(skip in lower for skip in ["pixel","tracking","beacon","logo","icon","sprite","1x1","blank"]):
            continue
        if re.search(r'\.(jpg|jpeg|webp|png|gif)', lower):
            return src
    return imgs[0] if imgs else ""


def extract_meta(html: str, name: str) -> str:
    m = re.search(r'<meta[^>]+name=["\']description["\'][^>]+content=["\'](.*?)["\']', html, re.I)
    if m and name == "description":
        return m.group(1)
    m2 = re.search(rf'<meta[^>]+property=["\']og:{name}["\'][^>]+content=["\'](.*?)["\']', html, re.I)
    if m2:
        return m2.group(1)
    return ""


def extract_h1(html: str) -> str:
    m = re.search(r'<h1[^>]*>(.*?)</h1>', html, re.I | re.S)
    if m:
        return re.sub(r'<[^>]+>', '', m.group(1)).strip()
    return ""


def extract_body_text(html: str) -> str:
    try:
        doc = lhtml.fromstring(html)
        for tag in doc.xpath('//script|//style|//noscript|//iframe|//head|//nav|//header|//footer'):
            parent = tag.getparent()
            if parent is not None:
                parent.remove(tag)
        container = (
            doc.xpath('//main') or
            doc.xpath('//article') or
            doc.xpath('//*[@id="content"]') or
            doc.xpath('//*[contains(concat(" ", @class, " "), " content ")]') or
            doc.xpath('//body')
        )
        node = container[0] if container else doc
        texts = node.xpath('.//text()')
        flat = ' '.join(t.strip() for t in texts if t.strip())
        flat = re.sub(r' {2,}', ' ', flat)
        return flat.strip()
    except Exception:
        text = re.sub(r'<[^>]+>', ' ', html)
        text = re.sub(r'\s+', ' ', text)
        return text.strip()


def extract_field(field_key: str, result) -> str:
    # If this result came from the Claude API, use pre-extracted fields directly
    if hasattr(result, "_extracted") and result._extracted:
        return result._extracted.get(field_key, "")

    html = result.html or ""
    meta = result.metadata or {}

    if field_key == "url":
        return result.url
    elif field_key == "title":
        return meta.get("title", "")
    elif field_key == "content":
        return extract_body_text(html)
    elif field_key == "image_url":
        return extract_hero_image(html)
    elif field_key == "meta_description":
        return extract_meta(html, "description")
    elif field_key == "h1":
        return extract_h1(html)
    elif field_key == "word_count":
        body = extract_body_text(html)
        return str(len(body.split())) if body else "0"
    elif field_key == "links":
        links = result.links or {}
        internal = [l.get("href", "") for l in links.get("internal", [])]
        return " | ".join(internal[:20])
    else:
        return ""


# ──────────────────────────────────────────────
# Excel builder
# ──────────────────────────────────────────────

FIELD_LABELS = {
    "url": "Page URL",
    "title": "Page Title",
    "content": "Main Content",
    "image_url": "Hero Image URL",
    "meta_description": "Meta Description",
    "h1": "H1 Heading",
    "word_count": "Word Count",
    "links": "Internal Links (first 20)",
}


def build_excel(rows: list, fields: list, filename: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crawl Results"

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=FIELD_LABELS.get(field, field))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=False, vertical="center")

    for row_idx, row_data in enumerate(rows, 2):
        fill = PatternFill("solid", fgColor="F8F8FC") if row_idx % 2 == 0 \
               else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, field in enumerate(fields, 1):
            val = row_data.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = {
        "url": 50, "title": 40, "content": 80, "image_url": 50,
        "meta_description": 60, "h1": 40, "word_count": 12, "links": 60,
    }
    for col_idx, field in enumerate(fields, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = \
            col_widths.get(field, 30)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ──────────────────────────────────────────────
# Async job runners
# ──────────────────────────────────────────────

async def run_sitemap_crawl(job_id: str, sitemap_url: str, fields: list,
                             filename: str, api_key: str, claude_api_url: str = ""):
    job = jobs[job_id]
    job["status"] = "fetching_sitemap"
    job["log"] = [f"Fetching sitemap: {sitemap_url}"]
    if api_key:
        job["log"].append("ScrapingBee fallback ENABLED (premium → stealth escalation)")
    else:
        job["log"].append("WARNING: No ScrapingBee API key — blocked pages will try Claude API only")
    if claude_api_url:
        job["log"].append("Claude API fallback ENABLED (tier 4 last resort)")

    try:
        async with aiohttp.ClientSession() as session:
            urls = await fetch_sitemap_urls(sitemap_url, session)
    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)
        return

    if not urls:
        job["status"] = "error"
        job["error"] = "No URLs found in sitemap."
        return

    job["total"] = len(urls)
    job["status"] = "crawling"
    job["log"].append(f"Found {len(urls)} URLs")
    if api_key:
        job["log"].append("ScrapingBee fallback enabled — will activate on blocked pages only")

    config = CrawlerRunConfig(
        cache_mode=CacheMode.BYPASS,
        word_count_threshold=10,
        page_timeout=30000,
    )

    rows = []
    semaphore = asyncio.Semaphore(5)

    async with AsyncWebCrawler() as crawler:
        async with aiohttp.ClientSession() as session:

            async def crawl_one(url):
                async with semaphore:
                    try:
                        result = await fetch_page(url, crawler, config, api_key, session, job, claude_api_url, fields)
                        if not result.success:
                            job["log"].append(f"SKIP {url}: {result.error_message}")
                            job["progress"] = job.get("progress", 0) + 1
                            return None
                        row = {f: extract_field(f, result) for f in fields}
                        job["progress"] = job.get("progress", 0) + 1
                        job["log"].append(f"OK {url}")
                        return row
                    except Exception as e:
                        job["progress"] = job.get("progress", 0) + 1
                        job["log"].append(f"ERR {url}: {e}")
                        return None

            tasks = [crawl_one(url) for url in urls]
            raw = await asyncio.gather(*tasks)
            rows = [r for r in raw if r is not None]

    job["status"] = "building_excel"
    excel_bytes = build_excel(rows, fields, filename)
    job["excel"] = excel_bytes
    job["filename"] = filename if filename.endswith(".xlsx") else filename + ".xlsx"
    job["status"] = "done"
    job["log"].append(f"Done. {len(rows)} pages crawled.")


async def run_spider_job(job_id: str, start_url: str, fields: list,
                          filename: str, max_pages: int, api_key: str, claude_api_url: str = ""):
    rows = await spider_crawl(job_id, start_url, fields, max_pages, api_key, claude_api_url)
    job = jobs[job_id]
    job["status"] = "building_excel"
    excel_bytes = build_excel(rows, fields, filename)
    job["excel"] = excel_bytes
    job["filename"] = filename if filename.endswith(".xlsx") else filename + ".xlsx"
    job["status"] = "done"
    job["log"].append(f"Done. {len(rows)} pages crawled.")


def run_async_job(job_id, mode, url, fields, filename, max_pages, api_key, claude_api_url):
    if mode == "sitemap":
        asyncio.run(run_sitemap_crawl(job_id, url, fields, filename, api_key, claude_api_url))
    else:
        asyncio.run(run_spider_job(job_id, url, fields, filename, max_pages, api_key, claude_api_url))


# ──────────────────────────────────────────────
# Flask routes
# ──────────────────────────────────────────────

@app.route("/")
def index():
    with open("index.html", "r", encoding="utf-8") as f:
        return app.response_class(
            response=f.read(),
            status=200,
            mimetype="text/html; charset=utf-8"
        )


@app.route("/api/crawl", methods=["POST"])
def start_crawl():
    data = request.json
    mode = data.get("mode", "sitemap")
    url = data.get("url", "").strip()
    filename = data.get("filename", "crawl_results").strip() or "crawl_results"
    fields = data.get("fields", ["url", "title", "content", "image_url"])
    max_pages = int(data.get("max_pages", 300))
    # Use server-side env var if set (deployed mode), otherwise accept from UI (local mode)
    api_key = os.environ.get("SCRAPINGBEE_API_KEY") or data.get("scrapingbee_key", "").strip()
    claude_api_url = os.environ.get("CLAUDE_API_URL") or data.get("claude_api_url", "").strip()

    if not url:
        return jsonify({"error": "URL is required"}), 400

    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "status": "queued",
        "progress": 0,
        "total": 0,
        "log": [],
        "excel": None,
        "filename": filename,
        "error": None,
    }

    t = threading.Thread(
        target=run_async_job,
        args=(job_id, mode, url, fields, filename, max_pages, api_key, claude_api_url),
        daemon=True
    )
    t.start()

    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
def job_status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "total": job["total"],
        "log": job["log"][-30:],   # UI only shows last 30; full log available via /api/log/<job_id>
        "error": job["error"],
        "filename": job["filename"],
    })


@app.route("/api/download/<job_id>")
def download(job_id):
    job = jobs.get(job_id)
    if not job or job["status"] != "done" or not job["excel"]:
        return jsonify({"error": "Not ready"}), 404
    buf = io.BytesIO(job["excel"])
    fname = job["filename"]
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/log/<job_id>")
def download_log(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    log_text = "\n".join(job["log"])
    buf = io.BytesIO(log_text.encode("utf-8"))
    fname = (job["filename"] or "crawl").replace(".xlsx", "") + "_log.txt"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="text/plain")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    host = "0.0.0.0"
    print(f"\n🕷  Crawler UI running at http://{host}:{port}\n")
    app.run(debug=False, host=host, port=port)
